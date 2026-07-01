import os
import io
import pandas as pd
import getpass
import openpyxl
import msoffcrypto
import spire.xls

# get_patient_ids.py
# Author: Sjors
# Description: Script to get patient IDs from codelist for selected RESP-nrs
#              and save to password-protected Excel file.
#              Required for getting data from SEIN

# Set dirs and parameters
data_dir = os.path.join("..", "data", "selection")
codelist_path = os.path.join("L:\\","her_knf_golf","Wetenschap","newtransport","0_EpiLab","4_RESPectDB","1_Codelist","20250327_RESPect_lijst_final_retrospectief+prospectief_incl_info - kopie.xlsx")

# Load codelist in read-only mode (!) + ask for password if needed
sheets_dict = {}

# Decrypt codelist
password = getpass.getpass("Please enter codelist password: ")
decrypted = io.BytesIO()
with open(codelist_path, "rb") as f:
    office_file = msoffcrypto.OfficeFile(f)
    office_file.load_key(password=password)
    office_file.decrypt(decrypted)
# Read all sheets into dictionary of DataFrames
for sheet_name in openpyxl.load_workbook(decrypted, read_only=True).sheetnames:
    skiprows = 1 if sheet_name == "RESPect_prospectief_digitaal_IC" else 0
    decrypted.seek(0)
    df = pd.read_excel(
        decrypted,
        sheet_name=sheet_name,
        dtype=str,
        engine="openpyxl",
        skiprows=skiprows,
        usecols=[0, 1]
    )
    sheets_dict[sheet_name] = df

# Load selected patient IDs
df_selected = pd.read_csv(os.path.join(data_dir, "selected_summary.csv"), dtype=str)
selected_ids = df_selected["Participant Id"].unique()

# Find and couple RESP-nrs with AZU-nrs
coupled_ids = []
for resp_id in selected_ids:
    found = False
    # Loop over sheets to find RESP_id
    for sheet_df in sheets_dict.values():
        # Search proper columns
        resp_col = "RESP_number" if "RESP_number" in sheet_df.columns else "RESP"
        azu_col = "AZU_number" if "AZU_number" in sheet_df.columns else "AZU_nummer"
        # Search for matching RESP_id
        match = sheet_df[sheet_df[resp_col] == resp_id]
        # If found, get AZU number and break
        if not match.empty:
            azu_number = match.iloc[0][azu_col]
            coupled_ids.append({"RESP_number": resp_id, "AZU_number": azu_number})
            found = True
            break
    # If not found, throw warning and add with None
    if not found:
        print(f"\033[93mWARNING:\033[0m {resp_id} not found in codelist.")
        coupled_ids.append({"RESP_number": resp_id, "AZU_number": None})
# Store in dataframe
coupled_df = pd.DataFrame(coupled_ids)
coupled_df = coupled_df.sort_values(by="RESP_number").reset_index(drop=True)

# Save to password-protected CSV (Excel file)
output_path = os.path.join(data_dir, "selected_patient_ids.xlsx")
password = getpass.getpass("Please enter password to encrypt output: ")
# Save DataFrame to Excel (temporary file)
temp_output_path = os.path.join(data_dir, "temp_selected_patient_ids.xlsx")
coupled_df.to_excel(temp_output_path, index=False)
# Encrypt Excel file
wb = spire.xls.Workbook()
wb.LoadFromFile(temp_output_path)
wb.Protect(password)
wb.SaveToFile(output_path, spire.xls.ExcelVersion.Version2013)
# Remove temporary file
os.remove(temp_output_path)

# Final message
print(f"Saved coupled patient IDs to {os.path.abspath(output_path)}")